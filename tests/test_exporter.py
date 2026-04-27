from openpyxl import load_workbook
import pandas as pd

from app.executor import ExecutionContext
from app.exporter import REQUIRED_SHEETS, export_workbook


def test_exporter_writes_required_sheets(tmp_path, sample_deal_pipeline):
    ctx = ExecutionContext(
        datasets={"deal_pipeline": sample_deal_pipeline},
        issues=pd.DataFrame([{"dataset": "deal_pipeline", "row_index": 1, "issue_type": "x", "column_name": "stage", "severity": "high", "detail": "bad"}]),
        api_enrichment=pd.DataFrame({"country": ["ID"], "region": ["East Asia & Pacific"]}),
        summary=pd.DataFrame({"owner": ["Ana"], "count": [1]}),
    )
    output = export_workbook(tmp_path / "out.xlsx", [{"dataset_name": "deal_pipeline", "row_count": 3, "column_count": 6, "duplicate_count": 0, "columns": ["company_name"], "basic_anomalies": []}], ctx)
    wb = load_workbook(output)
    assert set(REQUIRED_SHEETS).issubset(set(wb.sheetnames))

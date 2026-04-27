from openpyxl import load_workbook
import pandas as pd

from app.chatbot.action_parser import plan_from_selected_actions
from app.processing.executor import execute_cleaning_plan
from app.processing.excel_exporter import CHATBOT_OUTPUT_SHEETS
from app.processing.cleaners import fill_missing_values
from app.processing.profiler import profile_dataframe


def test_executor_waits_for_plan_and_creates_chatbot_workbook(tmp_path):
    df = pd.DataFrame(
        {
            "Deal Date": ["2026-01-01", "bad-date", "2026-01-01"],
            "Amount": ["$1,200", "2.5M", "one million"],
            "Name": [" Alpha ", " Alpha ", "Beta"],
        }
    )
    profile = profile_dataframe(df, "sample.csv", "csv")
    plan = plan_from_selected_actions(
        [
            "standardize_column_names",
            "trim_whitespace",
            "parse_date_columns",
            "convert_numeric_columns",
            "create_summary_sheet",
            "flag_invalid_rows",
            "create_inefficiency_report",
            "create_management_view",
            "enrich_with_api",
        ],
        profile,
    )
    output = tmp_path / "cleaned.xlsx"
    result = execute_cleaning_plan(df, profile, plan, output)
    assert output.exists()
    assert result.original_rows == 3
    assert result.invalid_dates_found >= 1
    assert result.numeric_conversion_issues == 1
    wb = load_workbook(output)
    assert set(CHATBOT_OUTPUT_SHEETS).issubset(set(wb.sheetnames))
    cleaned_sheet = wb["cleaned_data"]
    assert cleaned_sheet.freeze_panes == "A2"
    assert cleaned_sheet.auto_filter.ref is not None
    assert cleaned_sheet["A1"].font.bold is True
    assert cleaned_sheet["A1"].fill.fill_type == "solid"


def test_executor_rejects_unknown_action(tmp_path):
    from app.models import CleaningAction, CleaningPlan

    df = pd.DataFrame({"a": [1]})
    profile = profile_dataframe(df, "sample.csv", "csv")
    plan = CleaningPlan(actions=[CleaningAction(action_name="run_arbitrary_code")])
    try:
        execute_cleaning_plan(df, profile, plan, tmp_path / "out.xlsx")
    except ValueError as exc:
        assert "Unsupported cleaning action" in str(exc)
    else:
        raise AssertionError("Expected unknown action to be rejected")


def test_fill_missing_values_handles_nullable_integer_median_float():
    df = pd.DataFrame({"score": pd.Series([1, None, 2], dtype="Int64")})
    cleaned, issues, message = fill_missing_values(df)
    assert cleaned["score"].dtype == "float64"
    assert cleaned.loc[1, "score"] == 1.5
    assert len(issues) == 1
    assert "Filled 1 missing values" in message


def test_executor_sorts_rows_by_prompt_plan(tmp_path):
    from app.chatbot.action_parser import parse_user_instruction

    df = pd.DataFrame({"departemen": ["Sales", "Finance", "Ops"], "gaji": [700, 1200, 900]})
    profile = profile_dataframe(df, "sample.csv", "csv")
    parsed = parse_user_instruction("urutkan gaji dari terbesar", profile)
    assert parsed.plan is not None
    output = tmp_path / "sorted.xlsx"
    execute_cleaning_plan(df, profile, parsed.plan, output)
    sorted_df = pd.read_excel(output, sheet_name="cleaned_data")
    assert sorted_df["gaji"].tolist() == [1200, 900, 700]

from openpyxl import load_workbook
import pandas as pd

from app.excel_agent.executor import execute_plan
from app.excel_agent.profiler import profile_dataframe
from app.excel_agent.schemas import PlannerResult, ToolCall, WorkbookContext
from app.excel_agent.tools.registry import ToolContext


def test_split_sheet_by_column_creates_sheets(tmp_path):
    df = pd.DataFrame({"Divisi": ["Ops", "Finance"], "Sales": [100, 200]})
    profile = profile_dataframe(df, WorkbookContext(file_name="demo.csv", file_type="csv"))
    ctx = ToolContext(original_df=df, working_df=df.copy(), data_profile=profile)
    plan = PlannerResult(
        assistant_response="",
        tool_calls=[
            ToolCall(tool="split_sheet_by_column", args={"column": "Divisi"}),
            ToolCall(tool="export_workbook", args={"output_name": str(tmp_path / "split.xlsx")}),
        ],
    )
    execute_plan(ctx, plan)
    wb = load_workbook(tmp_path / "split.xlsx", read_only=True)
    assert "Divisi_Ops" in wb.sheetnames
    assert "Divisi_Finance" in wb.sheetnames

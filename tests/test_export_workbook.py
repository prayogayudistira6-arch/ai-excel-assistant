from openpyxl import load_workbook
import pandas as pd

from app.excel_agent.executor import execute_plan
from app.excel_agent.profiler import profile_dataframe
from app.excel_agent.schemas import PlannerResult, ToolCall, WorkbookContext
from app.excel_agent.tools.registry import ToolContext


def test_agent_export_workbook_minimal(tmp_path):
    df = pd.DataFrame({"A": [1]})
    profile = profile_dataframe(df, WorkbookContext(file_name="demo.csv", file_type="csv"))
    ctx = ToolContext(original_df=df, working_df=df.copy(), data_profile=profile)
    execute_plan(ctx, PlannerResult(assistant_response="", tool_calls=[ToolCall(tool="export_workbook", args={"output_name": str(tmp_path / "agent.xlsx")})]))
    wb = load_workbook(tmp_path / "agent.xlsx", read_only=True)
    assert {"original_preview", "cleaned_data", "data_profile", "operation_history"}.issubset(set(wb.sheetnames))

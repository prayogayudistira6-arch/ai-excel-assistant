from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.excel_agent.tools.registry import ToolContext


COLOR_MAP = {
    "red": "FECACA",
    "yellow": "FEF3C7",
    "green": "DCFCE7",
    "blue": "DBEAFE",
    "orange": "FED7AA",
    "purple": "E9D5FF",
    "gray": "E5E7EB",
}


def _profile_frame(ctx: ToolContext) -> pd.DataFrame:
    profile = ctx.data_profile
    if profile is None:
        return pd.DataFrame()
    rows = []
    for key, value in profile.model_dump().items():
        if isinstance(value, (dict, list)):
            value = str(value)
        rows.append({"metric": key, "value": value})
    return pd.DataFrame(rows)


def _flagged_issues(ctx: ToolContext) -> pd.DataFrame:
    frames = []
    for key in ["flagged_issues", "flagged_missing", "flagged_duplicates"]:
        value = ctx.artifacts.get(key)
        if isinstance(value, pd.DataFrame) and not value.empty:
            frames.append(value)
    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame(columns=["row_index", "column", "issue_type", "original_value", "suggested_fix"])


def _style(path: Path, formatting: list[dict[str, object]]) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2937")
    missing_fill = PatternFill("solid", fgColor="FEF3C7")
    issue_fill = PatternFill("solid", fgColor="FECACA")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 45)
        if ws.title == "cleaned_data" and ws.max_row > 1:
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    FormulaRule(formula=[f'ISBLANK({letter}2)'], fill=missing_fill),
                )
        if ws.title == "flagged_issues" and ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                for cell in ws[row]:
                    cell.fill = issue_fill
    if "cleaned_data" in wb.sheetnames:
        ws = wb["cleaned_data"]
        headers = [str(cell.value) for cell in ws[1]]
        for item in formatting:
            fill = PatternFill("solid", fgColor=COLOR_MAP.get(str(item.get("color", "yellow")), "FEF3C7"))
            if item.get("type") == "column" and item.get("column") in headers:
                col_idx = headers.index(str(item["column"])) + 1
                for row in range(1, ws.max_row + 1):
                    ws.cell(row, col_idx).fill = fill
            elif item.get("type") == "row_condition" and item.get("column") in headers:
                col_idx = headers.index(str(item["column"])) + 1
                expected = str(item.get("equals", "")).lower()
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row, col_idx).value).lower() == expected:
                        for cell in ws[row]:
                            cell.fill = fill
    wb.save(path)


def export_agent_workbook(output_path: str | Path, ctx: ToolContext) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    sheets: dict[str, pd.DataFrame] = {
        "original_preview": ctx.original_df.head(100),
        "cleaned_data": ctx.working_df,
        "flagged_issues": _flagged_issues(ctx),
        "data_profile": _profile_frame(ctx),
        "operation_history": pd.DataFrame({"operation": ctx.operation_history}),
    }
    for key, sheet_name in [
        ("management_report", "management_report"),
        ("group_summary", "group_summary"),
        ("pivot_table", "pivot_table"),
        ("api_enrichment", "api_enrichment"),
    ]:
        value = ctx.artifacts.get(key)
        if isinstance(value, pd.DataFrame):
            sheets[sheet_name] = value
    split_sheets = ctx.artifacts.get("split_sheets", {})
    if isinstance(split_sheets, dict):
        for name, frame in split_sheets.items():
            if isinstance(frame, pd.DataFrame):
                sheets[str(name)[:31]] = frame
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            safe_name = name[:31] or "sheet"
            frame.to_excel(writer, sheet_name=safe_name, index=False)
    _style(path, ctx.formatting)
    return path

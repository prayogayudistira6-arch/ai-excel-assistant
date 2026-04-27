from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill

from app.executor import ExecutionContext


REQUIRED_SHEETS = ["upload_profile", "cleaned_data", "flagged_issues", "api_enrichment", "summary", "action_log"]


def _profile_frame(profile: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for item in profile:
        rows.append({
            "dataset_name": item.get("dataset_name"),
            "row_count": item.get("row_count"),
            "column_count": item.get("column_count"),
            "duplicate_count": item.get("duplicate_count"),
            "columns": ", ".join(item.get("columns", [])),
            "basic_anomalies": "; ".join(item.get("basic_anomalies", [])),
        })
    return pd.DataFrame(rows)


def _style_workbook(path: str | Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2937")
    high_fill = PatternFill("solid", fgColor="FECACA")
    blank_fill = PatternFill("solid", fgColor="FEF3C7")
    for ws in wb.worksheets:
        if ws.max_row == 0:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for column in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_len + 2, 40)
        if ws.title == "flagged_issues" and ws.max_row > 1:
            headers = [cell.value for cell in ws[1]]
            if "severity" in headers:
                severity_col = headers.index("severity") + 1
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row, severity_col).value == "high":
                        for cell in ws[row]:
                            cell.fill = high_fill
        if ws.title == "cleaned_data" and ws.max_row > 1:
            for col in range(1, ws.max_column + 1):
                letter = ws.cell(1, col).column_letter
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    FormulaRule(formula=[f'ISBLANK({letter}2)'], fill=blank_fill),
                )
        if ws.title == "summary" and ws.max_row > 1 and ws.max_column > 0:
            headers = [cell.value for cell in ws[1]]
            if "count" in headers:
                letter = ws.cell(1, headers.index("count") + 1).column_letter
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    ColorScaleRule(start_type="min", start_color="DBEAFE", end_type="max", end_color="1D4ED8"),
                )
    wb.save(path)


def export_workbook(output_path: str | Path, profile: list[dict[str, Any]], ctx: ExecutionContext) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    cleaned = ctx.datasets["deal_pipeline"] if "deal_pipeline" in ctx.datasets else next(iter(ctx.datasets.values()), pd.DataFrame())
    action_log = pd.DataFrame([log.model_dump() for log in ctx.action_log])
    sheets = {
        "upload_profile": _profile_frame(profile),
        "cleaned_data": cleaned,
        "flagged_issues": ctx.issues,
        "api_enrichment": ctx.api_enrichment,
        "summary": ctx.summary,
        "action_log": action_log,
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name in REQUIRED_SHEETS:
            frame = sheets.get(name, pd.DataFrame())
            frame.to_excel(writer, sheet_name=name, index=False)
    _style_workbook(path)
    return path

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill


CHATBOT_OUTPUT_SHEETS = [
    "original_preview",
    "data_profile",
    "cleaned_data",
    "flagged_issues",
    "inefficiency_report",
    "ai_recommendations",
    "cleaning_plan",
    "api_enrichment",
    "management_view",
    "summary",
    "change_log",
]


def _summary_with_metrics(summary: pd.DataFrame, metrics: dict[str, Any]) -> pd.DataFrame:
    metric_rows = [{"metric": key, "value": value} for key, value in metrics.items()]
    if summary.empty:
        return pd.DataFrame(metric_rows)
    return pd.concat([pd.DataFrame(metric_rows), summary], ignore_index=True)


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fills = {
        "original_preview": "334155",
        "data_profile": "0F766E",
        "cleaned_data": "166534",
        "flagged_issues": "991B1B",
        "inefficiency_report": "B45309",
        "ai_recommendations": "4338CA",
        "cleaning_plan": "6D28D9",
        "api_enrichment": "0369A1",
        "management_view": "15803D",
        "summary": "1D4ED8",
        "change_log": "7C2D12",
    }
    missing_fill = PatternFill("solid", fgColor="FEF3C7")
    issue_fill = PatternFill("solid", fgColor="FECACA")
    valid_fill = PatternFill("solid", fgColor="DCFCE7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor=fills.get(ws.title, "1F2937"))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)

        for column in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 10), 45)

        if ws.title == "cleaned_data" and ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                if all(ws.cell(row, col).value not in {None, ""} for col in range(1, ws.max_column + 1)):
                    for cell in ws[row]:
                        cell.fill = valid_fill
            for col in range(1, ws.max_column + 1):
                letter = ws.cell(1, col).column_letter
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    FormulaRule(formula=[f'ISBLANK({letter}2)'], fill=missing_fill),
                )

        if ws.title == "flagged_issues" and ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                issue_type = str(ws.cell(row, 3).value or "")
                if issue_type in {"invalid_date", "invalid_numeric", "duplicate_row", "duplicate_by_columns"}:
                    for cell in ws[row]:
                        cell.fill = issue_fill
        if ws.title in {"inefficiency_report", "management_view", "ai_recommendations"} and ws.max_row > 1:
            headers = [cell.value for cell in ws[1]]
            severity_col = headers.index("severity") + 1 if "severity" in headers else None
            priority_col = headers.index("priority") + 1 if "priority" in headers else None
            target_col = severity_col or priority_col
            if target_col:
                for row in range(2, ws.max_row + 1):
                    value = str(ws.cell(row, target_col).value or "").lower()
                    if value == "high":
                        for cell in ws[row]:
                            cell.fill = issue_fill
                    elif value in {"medium", "warning"}:
                        for cell in ws[row]:
                            cell.fill = missing_fill
    wb.save(path)


def export_cleaning_workbook(
    output_path: str | Path,
    original_df: pd.DataFrame,
    cleaned_df: pd.DataFrame,
    flagged_issues: pd.DataFrame,
    summary: pd.DataFrame,
    data_profile: pd.DataFrame,
    inefficiency_report: pd.DataFrame,
    ai_recommendations: pd.DataFrame,
    cleaning_plan: pd.DataFrame,
    api_enrichment: pd.DataFrame,
    management_view: pd.DataFrame,
    change_log: list[str],
    metrics: dict[str, Any],
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    summary_df = _summary_with_metrics(summary, metrics)
    change_log_df = pd.DataFrame({"step": range(1, len(change_log) + 1), "change": change_log})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        sheets = {
            "original_preview": original_df.head(100),
            "data_profile": data_profile,
            "cleaned_data": cleaned_df,
            "flagged_issues": flagged_issues,
            "inefficiency_report": inefficiency_report,
            "ai_recommendations": ai_recommendations,
            "cleaning_plan": cleaning_plan,
            "api_enrichment": api_enrichment,
            "management_view": management_view,
            "summary": summary_df,
            "change_log": change_log_df,
        }
        for sheet_name in CHATBOT_OUTPUT_SHEETS:
            sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)

    _style_workbook(path)
    return path

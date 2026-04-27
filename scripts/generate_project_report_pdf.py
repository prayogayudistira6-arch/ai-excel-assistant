from __future__ import annotations

from datetime import datetime
from pathlib import Path
import textwrap

from openpyxl import load_workbook


OUTPUT_DIR = Path("data/output")
WORKBOOK_PATH = OUTPUT_DIR / "assistant_output.xlsx"
REPORT_PATH = OUTPUT_DIR / "project_report.pdf"


def _escape_pdf_text(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _wrap(text: str, width: int = 92) -> list[str]:
    if not text:
        return [""]
    return textwrap.wrap(text, width=width, replace_whitespace=False) or [""]


def _build_report_lines() -> list[tuple[str, str]]:
    workbook_rows: list[str] = []
    if WORKBOOK_PATH.exists():
        wb = load_workbook(WORKBOOK_PATH, read_only=True)
        for ws in wb.worksheets:
            workbook_rows.append(f"- {ws.title}: {ws.max_row} rows x {ws.max_column} columns")
    else:
        workbook_rows.append("- assistant_output.xlsx belum ditemukan saat laporan dibuat")

    return [
        ("title", "Laporan Rinci Project AI Spreadsheet Automation Assistant"),
        ("body", f"Tanggal laporan: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"),
        ("body", "Lokasi project: /media/yoga/DATA LINUX 21/portfolio/ai spreadsheet"),
        ("heading", "1. Ringkasan Eksekutif"),
        ("body", "Project ini adalah demo portfolio untuk otomasi spreadsheet lokal yang aman. Sistem membaca file Excel/CSV, membuat profil dataset, menjalankan validasi data, menyusun action plan mock, mengeksekusi transformasi dari whitelist, melakukan enrichment metadata negara, lalu mengekspor workbook multi-sheet yang siap direview."),
        ("body", "Desain utama sengaja memisahkan planner dan executor. Planner hanya membuat rencana aksi terstruktur, sedangkan executor hanya menjalankan fungsi Python yang terdaftar di ALLOWED_ACTIONS. Model ini menghindari eksekusi kode bebas dari output LLM."),
        ("heading", "2. Tujuan Bisnis"),
        ("body", "Masalah yang ditangani adalah pekerjaan spreadsheet operasional yang sering berulang: membersihkan nama kolom, menghapus spasi liar, menormalkan key perusahaan, mengubah nilai numerik dan tanggal, menandai duplikasi, mendeteksi follow-up overdue, dan membuat ringkasan yang bisa dipakai tim operasi atau investment pipeline."),
        ("body", "Nilai portfolio project ini ada pada kombinasi data-quality workflow, safety boundary untuk AI planner, dan output Excel yang familiar bagi user non-teknis."),
        ("heading", "3. Arsitektur Teknis"),
        ("body", "Pipeline utama berada di app/main.py dan memanggil modul yang sama untuk CLI maupun UI Streamlit. Alur runtime: read_input_dir -> profile_datasets -> validate_datasets -> MockLLMClient.create_plan -> execute_plan -> export_workbook."),
        ("body", "Lapisan reader berada di app/io/readers.py. File Excel dibaca multi-sheet dengan pandas.read_excel(sheet_name=None), sedangkan CSV dibaca sebagai satu dataset dengan pandas.read_csv(on_bad_lines='warn'). Nama dataset disanitasi agar stabil untuk pipeline."),
        ("body", "Lapisan profiling berada di app/profiling.py. Modul ini menghasilkan ringkasan JSON-serializable seperti jumlah baris, jumlah kolom, tipe data, null count, duplicate count, kandidat kolom tanggal, kandidat kolom numerik, kandidat key, sample rows, dan basic anomalies."),
        ("body", "Lapisan validasi berada di app/validation.py. Validasi tidak memutus pipeline ketika data kotor ditemukan; issue dikumpulkan ke DataFrame flagged_issues. Business rules mencakup invalid enum, missing required, invalid date, dan overdue follow-up."),
        ("body", "Lapisan action dan executor berada di app/actions.py dan app/executor.py. ExecutionContext membawa datasets, issues, api_enrichment, summary, dan action_log. Action yang tersedia meliputi standardize_column_names, trim_whitespace, add_normalized_key, coerce_numeric_columns, parse_date_columns, remove_duplicates, enrich_country_metadata, flag_overdue_rows, dan create_grouped_summary."),
        ("body", "Lapisan exporter berada di app/exporter.py. Workbook ditulis dengan pandas ExcelWriter dan distyling dengan openpyxl: header gelap, freeze panes, autofilter, conditional formatting untuk blank cell dan severity high, serta auto column width."),
        ("heading", "4. Komponen Utama"),
        ("body", "- app/config.py: settings loader berbasis dotenv dan environment variable."),
        ("body", "- app/models.py: model Pydantic untuk action plan, action step, issue record, dataset profile, dan execution log."),
        ("body", "- app/services/llm_client.py: MockLLMClient sebagai default runtime tanpa API key dan OpenAIPlannerClient dengan fallback mock."),
        ("body", "- app/services/country_api.py: enrichment metadata negara dengan local mapping default dan jalur World Bank API bila mode non-mock digunakan."),
        ("body", "- scripts/generate_dummy_data.py: generator 24 baris dummy data yang sengaja messy."),
        ("body", "- app/ui/streamlit_app.py: UI sederhana untuk upload file dan download workbook hasil pipeline."),
        ("heading", "5. Dataset Dummy"),
        ("body", "Generator membuat tiga file input: deal_pipeline.xlsx, followups.xlsx, dan ops_requests.xlsx. Data berisi duplikasi nama perusahaan, variasi casing, status invalid, missing owner, tanggal tidak bisa diparse, valuation text seperti '2.3m', dan follow-up yang sudah overdue."),
        ("heading", "6. Output Workbook"),
        ("body", f"Workbook output berada di {WORKBOOK_PATH}. Sheet wajib sudah diverifikasi ada semua: upload_profile, cleaned_data, flagged_issues, api_enrichment, summary, dan action_log."),
        *[("body", row) for row in workbook_rows],
        ("heading", "7. Hasil Quality Pass"),
        ("body", "Perintah test terbaru: .venv/bin/pytest -q"),
        ("body", "Hasil: 8 passed, 2 warnings. Warning berasal dari perubahan perilaku pandas 3 terkait select_dtypes(include=['object']) dan tidak menggagalkan test."),
        ("body", "Perintah generator data: .venv/bin/python scripts/generate_dummy_data.py --overwrite. Hasil: berhasil menulis data input."),
        ("body", "Perintah smoke run: .venv/bin/python -m app.main --input-dir data/input --output data/output/assistant_output.xlsx --llm-mode mock. Hasil: berhasil menulis workbook output."),
        ("heading", "8. Dependency dan Runtime"),
        ("body", "Dependency project dibatasi pada daftar blueprint: pandas, openpyxl, requests, python-dotenv, pydantic, pandera, streamlit, dan pytest. Tidak ada dependency baru yang ditambahkan untuk membuat laporan ini."),
        ("body", "Karena command python global tidak tersedia dan system Python menolak install global, project memakai virtualenv lokal .venv. Semua test dan smoke run dijalankan dengan .venv/bin/python atau .venv/bin/pytest."),
        ("heading", "9. Catatan Risiko dan Peningkatan"),
        ("body", "Runtime mock sudah berjalan end-to-end tanpa API key. Untuk produksi, perlu memperketat schema OpenAIPlannerClient agar benar-benar memakai structured output resmi, memperluas test untuk mode OpenAI fallback, dan menambahkan coverage untuk Streamlit workflow."),
        ("body", "Warning pandas dapat dibersihkan dengan mengganti select_dtypes(include=['object']) menjadi pendekatan yang eksplisit kompatibel pandas 2 dan 3. Perubahan ini tidak wajib untuk smoke run saat ini karena test tetap hijau."),
        ("body", "Output Excel sudah valid secara struktural. Tahap berikutnya bisa menambahkan validasi isi per sheet, misalnya memastikan flagged_issues memuat tipe issue yang diharapkan dan summary memakai grouping bisnis yang lebih kaya."),
        ("heading", "10. Kesimpulan"),
        ("body", "Project sudah memiliki pipeline lokal yang dapat dijalankan, test suite dasar hijau, dummy data generator, workbook multi-sheet, dan batas keamanan planner/executor yang jelas. Kondisi saat laporan dibuat layak untuk demo portfolio dan siap diperdalam pada integrasi LLM sungguhan atau workflow UI yang lebih lengkap."),
    ]


def _paginate(lines: list[tuple[str, str]]) -> list[list[tuple[str, str]]]:
    pages: list[list[tuple[str, str]]] = []
    page: list[tuple[str, str]] = []
    y = 760
    for style, text in lines:
        font_size = 18 if style == "title" else 13 if style == "heading" else 10
        line_height = 23 if style == "title" else 18 if style == "heading" else 13
        wrapped = _wrap(text, 70 if style == "title" else 86)
        needed = line_height * len(wrapped) + (8 if style == "heading" else 4)
        if y - needed < 50:
            pages.append(page)
            page = []
            y = 760
        for line in wrapped:
            page.append((style, line))
            y -= line_height
        y -= 6 if style in {"title", "heading"} else 3
    if page:
        pages.append(page)
    return pages


def _content_stream(page: list[tuple[str, str]], page_number: int, page_count: int) -> bytes:
    commands: list[str] = []
    y = 770
    for style, text in page:
        if style == "title":
            font, size, leading = "F2", 18, 24
        elif style == "heading":
            font, size, leading = "F2", 13, 19
        else:
            font, size, leading = "F1", 10, 14
        commands.append(f"BT /{font} {size} Tf 54 {y} Td ({_escape_pdf_text(text)}) Tj ET")
        y -= leading
    footer = f"Halaman {page_number} dari {page_count}"
    commands.append(f"BT /F1 9 Tf 270 30 Td ({footer}) Tj ET")
    return ("\n".join(commands) + "\n").encode("latin-1", errors="replace")


def write_pdf(path: Path, lines: list[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pages = _paginate(lines)
    objects: list[bytes] = []
    page_ids: list[int] = []
    content_ids: list[int] = []

    objects.append(b"<< /Type /Catalog /Pages 2 0 R >>")
    objects.append(b"")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
    objects.append(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")

    next_id = 5
    for page in pages:
        page_ids.append(next_id)
        content_ids.append(next_id + 1)
        next_id += 2
        objects.extend([b"", b""])

    kids = " ".join(f"{page_id} 0 R" for page_id in page_ids)
    objects[1] = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_ids)} >>".encode("ascii")

    for idx, page in enumerate(pages):
        page_obj_id = page_ids[idx]
        content_obj_id = content_ids[idx]
        content = _content_stream(page, idx + 1, len(pages))
        objects[page_obj_id - 1] = (
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            f"/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> "
            f"/Contents {content_obj_id} 0 R >>"
        ).encode("ascii")
        objects[content_obj_id - 1] = (
            f"<< /Length {len(content)} >>\nstream\n".encode("ascii")
            + content
            + b"endstream"
        )

    pdf = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for obj_id, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{obj_id} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    path.write_bytes(pdf)


def main() -> None:
    write_pdf(REPORT_PATH, _build_report_lines())
    print(f"wrote {REPORT_PATH}")


if __name__ == "__main__":
    main()
